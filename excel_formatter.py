import csv
import openpyxl as o
import os
import pandas as pd
import shutil
import subprocess
import sqlite3
import time
import chardet

def convert_csv_skip_rows_soffice(in_file, rows, out_file):
        """use soffice to convert xlsx to csv"""
        if os.name == 'nt':
            subprocess.run([r"C:\Program Files\LibreOffice\program\soffice.exe", "--headless", "--convert-to", "csv", in_file])
            csv_file = os.path.basename(in_file.replace(".xlsx", ".csv"))
            
            # read file
            with open(csv_file, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)  # store all rows in a list

            # skip the specified number of rows 
            data = data[rows:]

            # write the modified data back to file
            with open(out_file, 'w', newline='') as f_out:
                writer = csv.writer(f_out)
                writer.writerows(data)  # Write all remaining rows
        elif os.name == 'posix':
            subprocess.run(["soffice", "--headless", "--convert-to", "csv", in_file])
            
            csv_file = os.path.basename(in_file.replace(".xlsx", ".csv"))
            
            # read file
            with open(csv_file, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)  # store all rows in a list

            # skip the specified number of rows 
            data = data[rows:]

            # write the modified data back to file
            with open(out_file, 'w', newline='') as f_out:
                writer = csv.writer(f_out)
                writer.writerows(data)  # Write all remaining rows


def standard_order(file_path):
    """Formats standard order file and returns the processed file path."""
    start_time = time.time()  # Start timing
    
    # appending appropriate suffixes for transition files
    excel_transition_file = 'transition.xlsx'
    output_file = f'data/processed/{os.path.basename(file_path.replace(".xlsx", "_cleaned.xlsx"))}'

    # Convert xlsx to csv using win32
    file_csv = os.path.basename(file_path.replace(".xlsx", ".csv"))
    file_csv = os.path.abspath(file_csv)
    
    # conditional csv conversion based on operating system
    convert_csv_skip_rows_soffice(file_path, 7, file_csv)

    # load csv into pandas
    with open(file_csv, 'rb') as f:
        result = chardet.detect(f.read())
   
    df = pd.read_csv(file_csv, encoding=result['encoding'])
        
    df = df.astype(str)  # convert all data to string to enable SQL filtering

    # sql query to filter relevant columns 
    query = 'SELECT "Service O", "Order Pla", "Order Shi" FROM my_table'

    # filter using sqlite
    conn = sqlite3.connect(":memory:")  # use an in-memory database
    df.to_sql("my_table", conn, if_exists="replace", index=False)
    filtered_df = pd.read_sql(query, conn)

    # save filtered data to an excel file
    filtered_df.to_excel(excel_transition_file, index=False)

    # open the workbook for formatting
    wb = o.load_workbook(excel_transition_file)
    ws = wb.active

    # increase font size
    data_font = o.styles.Font(size=20)
    for row in ws:
        for cell in row:
            cell.font = data_font

    # increase column widths
    for col in ws.columns:
        max_length = 35
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)

    # clean up temporary files
    os.remove(file_csv)
    os.remove(excel_transition_file)

    # Calculate execution time
    execution_time = time.time() - start_time
    print(f"Standard order processing took {execution_time:.2f} seconds")

    # returns file path
    return output_file

def inventory_inbound(file_path):
    """Formats standard order file and returns the processed file path."""
    start_time = time.time()  # Start timing

    # appending appropriate suffixes for transition files
    excel_transition_file = 'transition.xlsx'
    output_file = f'data/processed/{os.path.basename(file_path.replace(".xlsx", "_cleaned.xlsx"))}'
    # Convert xlsx to csv using win32
    file_csv = os.path.basename(file_path.replace(".xlsx", ".csv"))
    file_csv = os.path.abspath(file_csv)

    # conditional csv conversion based on operating system
    convert_csv_skip_rows_soffice(file_path, 4, file_csv)
    
    # load csv into pandas
    with open(file_csv, 'rb') as f:
        result = chardet.detect(f.read())
   
    df = pd.read_csv(file_csv, encoding=result['encoding'])
    
    df = df.dropna(how='all') # drop rows with no values
    df = df.astype(str)  # convert all data to string to enable SQL filtering

    # sql query to filter relevant columns 
    query = 'SELECT "Trans Id", "Dock Date/Time Stamp", "Inbound Date" FROM my_table'

    # filter using sqlite
    conn = sqlite3.connect(":memory:")  # use an in-memory database
    df.to_sql("my_table", conn, if_exists="replace", index=False)
    filtered_df = pd.read_sql(query, conn)

    # save filtered data to an excel file
    filtered_df.to_excel(excel_transition_file, index=False)

    # open the workbook for formatting
    wb = o.load_workbook(excel_transition_file)
    ws = wb.active

    # increase font size
    data_font = o.styles.Font(size=20)
    for row in ws:
        for cell in row:
            cell.font = data_font

    # increase column widths
    for col in ws.columns:
        max_length = 35
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # change first column to numberic dtype
    col_letter = o.utils.get_column_letter(1)
    ws.column_dimensions[col_letter].number_format = 'General'
    for cell in ws[col_letter][1:]:
        if cell.value is not None:  # Skip header or empty cells
            cell.value = int(float(cell.value))

    # save the final cleaned file
    wb.save(output_file)

    # clean up temporary files
    os.remove(file_csv)
    os.remove(excel_transition_file)

    # Calculate execution time
    execution_time = time.time() - start_time
    print(f"Inventory inbound processing took {execution_time:.2f} seconds")

    # returns file path
    return output_file
